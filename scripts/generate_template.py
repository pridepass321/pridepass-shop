"""Generate batch orders Excel template with dropdowns, validation, and reference sheets."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from identities_data import HUE_PRESETS, PRONOUNS, all_identities, identity_labels

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "templates" / "lgbtiqasb-orders-template.xlsx"
OUT_CSV = ROOT / "templates" / "lgbtiqasb-orders-template.csv"

HEADER_FILL = PatternFill("solid", fgColor="1E1B4B")
ROW_ALT = PatternFill("solid", fgColor="0F0F18")
WHITE = Font(color="FFFFFF", bold=True)
THIN = Border(
    left=Side(style="thin", color="4C1D95"),
    right=Side(style="thin", color="4C1D95"),
    top=Side(style="thin", color="4C1D95"),
    bottom=Side(style="thin", color="4C1D95"),
)


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def autosize(ws, max_col, min_width=12, max_width=48):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        longest = min_width
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    longest = max(longest, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = longest


def build_lists_sheet(wb):
    ws = wb.create_sheet("_Lists")
    ws.sheet_state = "hidden"

    ws.cell(1, 1, "identity")
    for i, label in enumerate(identity_labels(), 2):
        ws.cell(i, 1, label)

    ws.cell(1, 2, "pronouns")
    for i, p in enumerate(PRONOUNS, 2):
        ws.cell(i, 2, p)

    ws.cell(1, 3, "hue_preset")
    for i, (name, hue, _sat) in enumerate(HUE_PRESETS, 2):
        ws.cell(i, 3, f"{name} ({hue})")

    last_id = len(identity_labels()) + 1
    last_pro = len(PRONOUNS) + 1
    return last_id, last_pro


def build_orders_sheet(wb, last_id, last_pro):
    ws = wb.active
    ws.title = "Orders"

    headers = [
        "order_id", "name", "identity", "pronouns",
        "member_number", "community_since", "hue", "saturation", "photo", "notes",
    ]
    descriptions = [
        "Optional reference", "REQUIRED — full name", "REQUIRED — pick from dropdown",
        "Optional — pick from dropdown", "Core cards only", "Identity cards only",
        "0–360 colour shift", "50–150 intensity", "Photo filename", "Not printed",
    ]

    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for i, d in enumerate(descriptions, 1):
        c = ws.cell(2, i, d)
        c.font = Font(italic=True, color="64748B", size=9)
        c.alignment = Alignment(wrap_text=True)

    examples = [
        ("ORD-001", "Alex Rivera", "Omnisexual", "they/them", "", "2024", 0, 100, "alex.jpg", "Sample"),
        ("ORD-002", "Sam Chen", "Demigirl", "she/they", "", "2025", 45, 110, "sam.png", "Warm hue"),
        ("ORD-003", "Jordan Lee", "Pride", "he/him", "LGB-2026-0001", "", 0, 100, "", "Core card"),
        ("ORD-004", "Morgan Blake", "Aceflux", "xe/xem", "", "2023", 180, 110, "", "Ocean hue"),
        ("ORD-005", "Riley Torres", "Quoiromantic / WTFromantic", "any pronouns", "", "2022", 90, 105, "riley.jpg", ""),
        ("", "", "", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", "", "", ""),
    ]
    for r, row in enumerate(examples, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            if r % 2 == 0:
                cell.fill = ROW_ALT

    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 36
    autosize(ws, len(headers))

    # Dropdown: identity (column C)
    dv_id = DataValidation(
        type="list",
        formula1=f"=_Lists!$A$2:$A${last_id}",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid identity",
        error="Choose a value from the Valid Identities list.",
    )
    dv_id.add("C3:C500")
    ws.add_data_validation(dv_id)

    # Dropdown: pronouns (column D)
    dv_pro = DataValidation(
        type="list",
        formula1=f"=_Lists!$B$2:$B${last_pro}",
        allow_blank=True,
    )
    dv_pro.add("D3:D500")
    ws.add_data_validation(dv_pro)

    # Hue 0-360
    dv_hue = DataValidation(type="whole", operator="between", formula1="0", formula2="360", allow_blank=True)
    dv_hue.add("G3:G500")
    ws.add_data_validation(dv_hue)

    # Saturation 50-150
    dv_sat = DataValidation(type="whole", operator="between", formula1="50", formula2="150", allow_blank=True)
    dv_sat.add("H3:H500")
    ws.add_data_validation(dv_sat)


def build_identities_sheet(wb):
    ws = wb.create_sheet("Valid Identities")
    ws.append(["identity", "category", "second_field", "card_file", "notes"])
    style_header_row(ws, 1, 5)

    for identity in all_identities():
        field = "member_number" if identity.get("fieldType") == "member" else "community_since"
        ws.append([
            identity["label"],
            identity.get("category", ""),
            field,
            f"{identity['id']}.png",
            identity.get("meaning", "Core community card"),
        ])

    ws.freeze_panes = "A2"
    autosize(ws, 5)


def build_instructions_sheet(wb):
    ws = wb.create_sheet("Instructions")
    lines = [
        ("LGBTIQASB+ Batch Print Template", True),
        ("", False),
        ("ORDERS SHEET", True),
        ("• One row per card. Rows 3+ are your orders.", False),
        ("• Required: name + identity (use dropdowns).", False),
        ("• Core cards → member_number. Identity cards → community_since.", False),
        ("• hue (0–360) shifts card colours. saturation (50–150) adjusts intensity.", False),
        ("• photo = filename only. Put photos in a folder and use --photos-dir when printing.", False),
        ("", False),
        ("PYTHON BATCH PRINT (no browser needed)", True),
        ("  python scripts/batch_print.py templates/lgbtiqasb-orders-template.xlsx", False),
        ("  python scripts/batch_print.py orders.xlsx --photos-dir ./photos --output print.pdf", False),
        ("", False),
        ("HUE PRESETS", True),
    ]
    for name, hue, _ in HUE_PRESETS:
        lines.append((f"  {hue:>3}  {name}", False))

    lines += [
        ("", False),
        ("PRINT", True),
        ("• Output PDF is CR80 (85.6 × 53.98 mm) — Evolis Primacy 2 ready.", False),
        ("• Or use the website Batch Print tab in your browser.", False),
    ]

    for r, (text, bold) in enumerate(lines, 1):
        cell = ws.cell(r, 1, text)
        if bold:
            cell.font = Font(bold=True, size=12 if r == 1 else 11, color="C4B5FD")
    ws.column_dimensions["A"].width = 78


def build_csv():
    headers = ["order_id", "name", "identity", "pronouns", "member_number", "community_since", "hue", "saturation", "photo", "notes"]
    desc = [
        "Optional", "REQUIRED", "REQUIRED — see Valid Identities", "Optional",
        "Core cards only", "Identity cards only", "0-360", "50-150", "Photo filename", "Not printed",
    ]
    examples = [
        ["ORD-001", "Alex Rivera", "Omnisexual", "they/them", "", "2024", "0", "100", "alex.jpg", ""],
        ["ORD-002", "Sam Chen", "Demigirl", "she/they", "", "2025", "45", "110", "sam.png", ""],
        ["ORD-003", "Jordan Lee", "Pride", "he/him", "LGB-2026-0001", "", "0", "100", "", ""],
    ]
    lines = [",".join(f'"{c}"' for c in row) for row in [headers, desc, *examples]]
    OUT_CSV.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Created {OUT_CSV}")


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    last_id, last_pro = build_lists_sheet(wb)
    build_orders_sheet(wb, last_id, last_pro)
    build_identities_sheet(wb)
    build_instructions_sheet(wb)
    wb.save(OUT)
    print(f"Created {OUT}")
    build_csv()


if __name__ == "__main__":
    main()