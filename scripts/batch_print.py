#!/usr/bin/env python3
"""
Batch-print community access cards from spreadsheet — no browser required.

Usage:
  python scripts/batch_print.py templates/lgbtiqasb-orders-template.xlsx
  python scripts/batch_print.py orders.xlsx --photos-dir ./photos --output batch.pdf
  python scripts/batch_print.py orders.xlsx --no-back --output cards/
"""
from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas as pdf_canvas

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

from card_render import render_card_back, render_card_front
from identities_data import resolve_identity

ROOT = SCRIPT_DIR.parent
CR80_W_MM = 85.6
CR80_H_MM = 53.98


def load_orders(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Orders" if "Orders" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h or "").strip().lower().replace(" ", "_") for h in rows[0]]
    alias = {
        "community-since": "community_since",
        "member-number": "member_number",
        "order-id": "order_id",
    }
    headers = [alias.get(h, h) for h in headers]

    orders = []
    for raw in rows[1:]:
        if not raw or not any(raw):
            continue
        row = {headers[i]: (raw[i] if i < len(raw) else "") for i in range(len(headers))}
        first = str(row.get(headers[0], "")).lower()
        if "required" in first or "optional" in first:
            continue
        orders.append(row)
    wb.close()
    return orders


def resolve_photo(photos_dir: Path | None, filename: str) -> Path | None:
    if not photos_dir or not filename:
        return None
    name = Path(str(filename).strip()).name
    candidate = photos_dir / name
    return candidate if candidate.exists() else None


def save_pdf(images: list, output: Path):
    c = pdf_canvas.Canvas(str(output), pagesize=(CR80_W_MM, CR80_H_MM))
    for img in images:
        import io
        buf = io.BytesIO()
        img.save(buf, format="PNG", optimize=True)
        buf.seek(0)
        c.drawImage(ImageReader(buf), 0, 0, width=CR80_W_MM, height=CR80_H_MM, preserveAspectRatio=True, anchor="c")
        c.showPage()
    c.save()


def main():
    parser = argparse.ArgumentParser(description="Batch print LGBTIQASB+ cards to PDF")
    parser.add_argument("spreadsheet", type=Path, help="Orders .xlsx or .csv path")
    parser.add_argument("--photos-dir", type=Path, default=None, help="Folder with order photos")
    parser.add_argument("--output", type=Path, default=None, help="Output PDF path or folder")
    parser.add_argument("--no-back", action="store_true", help="Skip card back pages")
    args = parser.parse_args()

    if not args.spreadsheet.exists():
        raise SystemExit(f"Spreadsheet not found: {args.spreadsheet}")

    if args.spreadsheet.suffix.lower() == ".csv":
        import csv
        with args.spreadsheet.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            orders = [dict(r) for r in reader if r.get("name")]
    else:
        orders = load_orders(args.spreadsheet)

    if not orders:
        raise SystemExit("No orders found in spreadsheet.")

    output = args.output or ROOT / "output" / f"lgbtiqasb-batch-{date.today().isoformat()}.pdf"
    output.parent.mkdir(parents=True, exist_ok=True)

    pages: list = []
    errors: list[str] = []

    for i, row in enumerate(orders, start=2):
        identity = resolve_identity(row.get("identity") or row.get("card") or row.get("design"))
        if not identity:
            errors.append(f"Row {i}: unknown identity '{row.get('identity')}'")
            continue

        photo = resolve_photo(args.photos_dir, str(row.get("photo") or row.get("photo_filename") or ""))
        try:
            front = render_card_front(
                identity_id=identity["id"],
                name=str(row.get("name") or row.get("full_name") or ""),
                member_number=str(row.get("member_number") or row.get("member") or ""),
                community_since=str(row.get("community_since") or row.get("since") or row.get("year") or "2026"),
                pronouns=str(row.get("pronouns") or ""),
                photo_path=photo,
                hue=float(row.get("hue") or 0),
                saturation=float(row.get("saturation") or 100),
            )
            pages.append(front)
            if not args.no_back:
                pages.append(render_card_back(front.width, front.height))
        except Exception as exc:
            errors.append(f"Row {i}: {exc}")

    if not pages:
        raise SystemExit("No cards generated.\n" + "\n".join(errors))

    if output.suffix.lower() == ".pdf":
        save_pdf(pages, output)
        print(f"Created {output} ({len(pages)} page(s))")
    else:
        output.mkdir(parents=True, exist_ok=True)
        for idx, page in enumerate(pages, 1):
            page.save(output / f"card-{idx:04d}.png", "PNG")
        print(f"Created {len(pages)} PNG(s) in {output}")

    if errors:
        print("\nWarnings:")
        for err in errors:
            print(f"  - {err}")


if __name__ == "__main__":
    main()